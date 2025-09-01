import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, LineChart, PieChart
from openpyxl.utils.dataframe import dataframe_to_rows

class Analysis:
    

    def get_errors_over_time(self,errors_df: pd.DataFrame, output_file: str = "errors_over_time.xlsx") -> pd.DataFrame:
        # Ensure timestamp is parsed and timezone dropped
        errors_df['time'] = pd.to_datetime(errors_df['time']).dt.tz_localize(None)

        # Group by month (year-month for multi-year safety)
        time_distribution = (
            errors_df.groupby(errors_df['time'].dt.to_period("M"))
            .size()
            .reset_index(name="error_count")
        )
        time_distribution['month'] = time_distribution['time'].astype(str)  # convert Period to str
        time_distribution = time_distribution[['month', 'error_count']]

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Errors Over Time"

        # Write DataFrame to Excel
        for r in dataframe_to_rows(time_distribution, index=False, header=True):
            ws.append(r)

        # Create line chart
        chart = LineChart()
        chart.title = "Errors Over Time (Monthly)"
        chart.y_axis.title = "Error Count"
        chart.x_axis.title = "Month"

        data = Reference(ws, min_col=2, min_row=1, max_row=len(time_distribution) + 1)  # error_count
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(time_distribution) + 1)  # month

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, "E2")

        # Save Excel
        wb.save(output_file)
        print(f"✅ Saved monthly errors graph to {output_file}")



    def analyze_balance_sync(self, transactions_df, errors_df) -> pd.DataFrame:
        # Step 1: Parse error messages to extract userId + balances
        sync_pattern = re.compile(
            r"userId:\s*'(?P<userId>[^']+)'.*?subscriptionBalance:\s*(?P<subscriptionBalance>\d+).*?paymentBalance:\s*(?P<paymentBalance>\d+)",
            re.DOTALL
        )

        error_records = []
        for _, row in errors_df.iterrows():
            match = sync_pattern.search(row['message'])
            if match:
                error_records.append({
                    "request_id": row['request_id'],
                    "userId": match.group("userId"),
                    "subscriptionBalance": int(match.group("subscriptionBalance")),
                    "paymentBalance": int(match.group("paymentBalance")),
                    "time": row["time"],   
                })

        errors_parsed_df = pd.DataFrame(error_records)

        if errors_parsed_df.empty:
            return pd.DataFrame(columns=[
                "userId", "total_transactions", "total_error_transactions",
                "total_debit_loss", "total_credit_loss", "currency", "first_error_time"
            ])

        # Step 2: Join with transactions on request_id
        merged_df = pd.merge(
            transactions_df,
            errors_parsed_df,
            on=["request_id", "userId"],
            how="inner"  # only keep those transactions linked to sync errors
        )

        # Step 3: Aggregate metrics
        grouped = transactions_df.groupby("userId").agg(
            total_transactions=("id", "count"),
            currency=("currency", "first")
        ).reset_index()

        error_stats = merged_df.groupby("userId").agg(
            total_error_transactions=("id", "count"),
            total_debit_loss=("amount", lambda x: x[merged_df.loc[x.index, "type"] == "DEBIT"].sum()),
            total_credit_loss=("amount", lambda x: x[merged_df.loc[x.index, "type"] == "CREDIT"].sum()),
            first_error_time=("time", "min"),   
        ).reset_index()

        # Step 4: Combine
        final_df = pd.merge(grouped, error_stats, on="userId", how="inner").fillna(0)

        return final_df

    def get_top_error_reasons(self, errors_df: pd.DataFrame, transactions_df: pd.DataFrame, output_file: str = "top_error_reasons_pie.xlsx") -> pd.DataFrame:
        # Join errors with transactions on request_id to get the action
        merged = errors_df.merge(
            transactions_df[['request_id', 'action']], 
            on="request_id", 
            how="left"
        )

        # Count by action
        reason_distribution = (
            merged.groupby("action")
            .size()
            .reset_index(name="error_count")
            .sort_values(by="error_count", ascending=False)
        )

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Top Error Reasons"

        # Write DataFrame to Excel
        for r in dataframe_to_rows(reason_distribution, index=False, header=True):
            ws.append(r)

        # Create pie chart
        pie = PieChart()
        pie.title = "Top Error Reasons Distribution"

        data = Reference(ws, min_col=2, min_row=1, max_row=len(reason_distribution) + 1)  # error_count
        labels = Reference(ws, min_col=1, min_row=2, max_row=len(reason_distribution) + 1)  # action

        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)

        ws.add_chart(pie, "E2")

        # Save Excel
        wb.save(output_file)
        print(f"✅ Saved top error reasons pie chart to {output_file}")
        
        

    def get_total_loss_by_currency(self, errors_df: pd.DataFrame, output_file: str = "loss_by_currency.xlsx") -> pd.DataFrame:
        # Aggregate debit and credit loss by currency
        loss_distribution = (
            errors_df.groupby("currency")[["total_debit_loss", "total_credit_loss"]]
            .sum()
            .reset_index()
        )

        # Save to Excel with a grouped bar chart
        wb = Workbook()
        ws = wb.active
        ws.title = "Loss by Currency"

        for r in dataframe_to_rows(loss_distribution, index=False, header=True):
            ws.append(r)

        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Total Debit & Credit Loss by Currency"
        chart.y_axis.title = "Loss Amount"
        chart.x_axis.title = "Currency"
        chart.grouping = "clustered"

        data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=len(loss_distribution) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(loss_distribution) + 1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, "E2")

        wb.save(output_file)
        print(f"✅ Saved loss chart to {output_file}")






